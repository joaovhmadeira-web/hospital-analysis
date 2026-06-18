"""
Geração de relatórios gerenciais nos formatos Excel e PDF.
"""
import io
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
import config
from database import engine

router = APIRouter(prefix="/api/relatorios", tags=["Relatórios Gerenciais"])

AZUL_HOSPITAL   = "1565C0"
TEAL_HOSPITAL   = "00838F"
CINZA_CABECALHO = "ECEFF1"

# ─────────────────────────────────────────────────────────────────────────────
# Helpers de estilo Excel
# ─────────────────────────────────────────────────────────────────────────────

def _cabecalho_excel(ws, titulo: str, subtitulo: str):
    ws.merge_cells("A1:H1")
    ws["A1"] = titulo
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=AZUL_HOSPITAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = subtitulo
    ws["A2"].font      = Font(italic=True, size=10, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    ws.append([])  # linha em branco


def _estilo_linha_cabecalho(ws, linha: int, ncols: int):
    fill = PatternFill("solid", fgColor=TEAL_HOSPITAL)
    font = Font(bold=True, color="FFFFFF")
    borda = Border(
        bottom=Side(style="medium", color=AZUL_HOSPITAL),
    )
    for col in range(1, ncols + 1):
        cell = ws.cell(row=linha, column=col)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = Alignment(horizontal="center")
        cell.border    = borda


def _ajustar_colunas(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 45)


# ─────────────────────────────────────────────────────────────────────────────
# Geração Excel
# ─────────────────────────────────────────────────────────────────────────────

def _gerar_excel(periodo_dias: int) -> io.BytesIO:
    wb = Workbook()
    agora = datetime.now()
    data_str = agora.strftime("%d/%m/%Y %H:%M")

    # ── Aba 1: Censo de Leitos ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Censo de Leitos"
    _cabecalho_excel(ws, "CENSO DE LEITOS — SITUAÇÃO ATUAL", f"Gerado em {data_str}")

    df = pd.read_sql(
        """
        SELECT tl.nome AS "Tipo", s.nome AS "Setor",
               COUNT(*) AS "Total",
               SUM(CASE WHEN l.status='disponivel' THEN 1 ELSE 0 END) AS "Disponíveis",
               SUM(CASE WHEN l.status='ocupado' THEN 1 ELSE 0 END)    AS "Ocupados",
               SUM(CASE WHEN l.status='manutencao' THEN 1 ELSE 0 END) AS "Manutenção",
               ROUND(SUM(CASE WHEN l.status='ocupado' THEN 1 ELSE 0 END)::numeric / COUNT(*) * 100, 1) AS "Taxa Ocupação (%)"
        FROM leitos l
        JOIN tipos_leito tl ON l.tipo_id = tl.id
        JOIN setores s ON l.setor_id = s.id
        GROUP BY tl.id, tl.nome, s.id, s.nome
        ORDER BY tl.id, s.nome
        """,
        engine,
    ).fillna(0)

    ws.append(list(df.columns))
    _estilo_linha_cabecalho(ws, ws.max_row, len(df.columns))
    for _, row in df.iterrows():
        ws.append(list(row))
    _ajustar_colunas(ws)

    # ── Aba 2: Internações ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Internações")
    _cabecalho_excel(ws2, f"INTERNAÇÕES — ÚLTIMOS {periodo_dias} DIAS", f"Gerado em {data_str}")

    df2 = pd.read_sql(
        f"""
        SELECT p.nome AS "Paciente",
               to_char(p.data_nascimento, 'DD/MM/YYYY') AS "Data Nasc.",
               p.sexo AS "Sexo",
               d.cid_codigo AS "CID",
               d.descricao  AS "Diagnóstico",
               l.numero AS "Leito",
               tl.nome  AS "Tipo Leito",
               to_char(i.data_entrada, 'DD/MM/YYYY HH24:MI') AS "Data Entrada",
               to_char(i.data_alta, 'DD/MM/YYYY HH24:MI') AS "Alta / Transferência",
               i.status AS "Status"
        FROM internacoes i
        JOIN pacientes   p  ON i.paciente_id = p.id
        JOIN diagnosticos d ON i.diagnostico_principal_id = d.id
        JOIN leitos l       ON i.leito_id = l.id
        JOIN tipos_leito tl ON l.tipo_id  = tl.id
        WHERE i.data_entrada >= NOW() - INTERVAL '{periodo_dias} days'
        ORDER BY i.data_entrada DESC
        """,
        engine,
    )

    ws2.append(list(df2.columns))
    _estilo_linha_cabecalho(ws2, ws2.max_row, len(df2.columns))
    for _, row in df2.iterrows():
        ws2.append(list(row))
    _ajustar_colunas(ws2)

    # ── Aba 3: Fila do PA ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Fila do PA")
    _cabecalho_excel(ws3, f"FLUXO DO PA — ÚLTIMOS {periodo_dias} DIAS", f"Gerado em {data_str}")

    df3 = pd.read_sql(
        f"""
        SELECT to_char(data_chegada, 'DD/MM/YYYY') AS "Data",
               COUNT(*) AS "Total Atendimentos",
               SUM(CASE WHEN status IN ('aguardando','em_atendimento') THEN 1 ELSE 0 END) AS "Aguardando",
               SUM(CASE WHEN status = 'atendido' THEN 1 ELSE 0 END)  AS "Atendidos",
               SUM(CASE WHEN status = 'desistiu' THEN 1 ELSE 0 END)  AS "Desistências",
               ROUND(AVG(CASE WHEN data_atendimento IS NOT NULL
                   THEN EXTRACT(EPOCH FROM (data_atendimento - data_chegada)) / 60 END), 0
               ) AS "T. Médio Espera (min)"
        FROM fila_espera
        WHERE data_chegada >= NOW() - INTERVAL '{periodo_dias} days'
        GROUP BY CAST(data_chegada AS DATE), to_char(data_chegada, 'DD/MM/YYYY')
        ORDER BY CAST(data_chegada AS DATE)
        """,
        engine,
    ).fillna(0)

    ws3.append(list(df3.columns))
    _estilo_linha_cabecalho(ws3, ws3.max_row, len(df3.columns))
    for _, row in df3.iterrows():
        ws3.append(list(row))
    _ajustar_colunas(ws3)

    # ── Aba 4: Estoque ────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Farmácia e ALMOX")
    _cabecalho_excel(ws4, "ESTOQUE — FARMÁCIA E ALMOXARIFADO", f"Gerado em {data_str}")

    df4 = pd.read_csv(config.DATA_DIR / "estoque_medicamentos.csv")
    df4 = df4.rename(columns={
        "nome": "Insumo", "categoria": "Categoria", "unidade": "Unidade",
        "quantidade_atual": "Qtd. Atual", "quantidade_minima": "Qtd. Mínima",
        "percentual_disponibilidade": "Disponib. (%)", "status_estoque": "Status",
        "fornecedor": "Fornecedor", "validade": "Validade",
        "localizacao_almoxarifado": "Localização",
    }).drop(columns=["id"])

    ws4.append(list(df4.columns))
    _estilo_linha_cabecalho(ws4, ws4.max_row, len(df4.columns))

    STATUS_CORES = {"esgotado": "FFCDD2", "critico": "FFE0B2", "baixo": "FFF9C4", "adequado": "E8F5E9"}
    for _, row in df4.iterrows():
        ws4.append(list(row))
        cor = STATUS_CORES.get(str(row["Status"]), "FFFFFF")
        for col in range(1, len(df4.columns) + 1):
            ws4.cell(row=ws4.max_row, column=col).fill = PatternFill("solid", fgColor=cor)
    _ajustar_colunas(ws4)

    # ── Aba 5: Escala ─────────────────────────────────────────────────────────
    ws5 = wb.create_sheet("Escala de Profissionais")
    _cabecalho_excel(ws5, "ESCALA DE PROFISSIONAIS", f"Gerado em {data_str}")

    df5 = pd.read_sql(
        """
        SELECT pro.nome AS "Profissional",
               pro.tipo AS "Tipo",
               COALESCE(e.nome,'—') AS "Especialidade",
               s.nome AS "Setor",
               p.data AS "Data",
               p.turno AS "Turno",
               to_char(p.inicio, 'HH24:MI') AS "Início",
               to_char(p.fim,    'HH24:MI') AS "Fim"
        FROM plantoes p
        JOIN profissionais pro ON p.profissional_id = pro.id
        JOIN setores s ON p.setor_id = s.id
        LEFT JOIN especialidades e ON pro.especialidade_id = e.id
        WHERE p.data >= CURRENT_DATE - INTERVAL '7 days'
        ORDER BY p.data, CASE p.turno WHEN 'manha' THEN 1 WHEN 'tarde' THEN 2 WHEN 'noite' THEN 3 END, s.nome
        """,
        engine,
    )

    ws5.append(list(df5.columns))
    _estilo_linha_cabecalho(ws5, ws5.max_row, len(df5.columns))
    for _, row in df5.iterrows():
        ws5.append(list(row))
    _ajustar_colunas(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# Geração PDF
# ─────────────────────────────────────────────────────────────────────────────

def _gerar_pdf(periodo_dias: int) -> io.BytesIO:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=2*cm, bottomMargin=2*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    styles = getSampleStyleSheet()
    azul   = rl_colors.HexColor(f"#{AZUL_HOSPITAL}")
    teal   = rl_colors.HexColor(f"#{TEAL_HOSPITAL}")
    agora  = datetime.now().strftime("%d/%m/%Y %H:%M")

    titulo_style = ParagraphStyle("Titulo",  parent=styles["Title"],
                                  textColor=azul, fontSize=18, spaceAfter=4)
    sub_style    = ParagraphStyle("Sub",     parent=styles["Normal"],
                                  textColor=rl_colors.grey, fontSize=9)
    secao_style  = ParagraphStyle("Secao",   parent=styles["Heading2"],
                                  textColor=teal, fontSize=12, spaceBefore=12, spaceAfter=4)

    def tabela_padrao(dados, cabecalho):
        rows = [cabecalho] + dados
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0,0), (-1,0), azul),
            ("TEXTCOLOR",   (0,0), (-1,0), rl_colors.white),
            ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",    (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [rl_colors.white, rl_colors.HexColor("#F1F5F9")]),
            ("GRID",        (0,0), (-1,-1), 0.3, rl_colors.HexColor("#CBD5E1")),
            ("VALIGN",      (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING",  (0,0), (-1,-1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
        ]))
        return t

    story = []

    # Capa
    story.append(Paragraph("RELATÓRIO GERENCIAL", titulo_style))
    story.append(Paragraph(f"Hospital Público — Gerado em {agora}", sub_style))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=azul))
    story.append(Spacer(1, 0.4*cm))

    # Seção 1: Censo de Leitos
    story.append(Paragraph("1. Censo de Leitos — Situação Atual", secao_style))
    df_leitos = pd.read_sql(
        """
        SELECT tl.nome AS "Tipo",
               COUNT(*) AS "Total",
               SUM(CASE WHEN l.status='disponivel' THEN 1 ELSE 0 END) AS "Disponíveis",
               SUM(CASE WHEN l.status='ocupado' THEN 1 ELSE 0 END)    AS "Ocupados",
               ROUND(SUM(CASE WHEN l.status='ocupado' THEN 1 ELSE 0 END)::numeric / COUNT(*) * 100, 1) AS "Taxa (%)"
        FROM leitos l
        JOIN tipos_leito tl ON l.tipo_id = tl.id
        GROUP BY tl.id, tl.nome ORDER BY tl.id
        """,
        engine,
    ).fillna(0)

    dados = [[str(v) for v in row] for _, row in df_leitos.iterrows()]
    story.append(tabela_padrao(dados, list(df_leitos.columns)))
    story.append(Spacer(1, 0.4*cm))

    # Seção 2: Internações ativas
    story.append(Paragraph("2. Internações Ativas por Diagnóstico", secao_style))
    df_int = pd.read_sql(
        """
        SELECT d.cid_codigo AS "CID",
               d.descricao  AS "Diagnóstico",
               COUNT(*)     AS "Internações"
        FROM internacoes i
        JOIN diagnosticos d ON i.diagnostico_principal_id = d.id
        WHERE i.status = 'ativa'
        GROUP BY d.id, d.cid_codigo, d.descricao ORDER BY 3 DESC LIMIT 12
        """,
        engine,
    )
    dados = [[str(v) for v in row] for _, row in df_int.iterrows()]
    story.append(tabela_padrao(dados, list(df_int.columns)))
    story.append(Spacer(1, 0.4*cm))

    # Seção 3: PA - últimos 7 dias
    story.append(Paragraph(f"3. Fluxo do PA — Últimos {periodo_dias} Dias", secao_style))
    df_pa = pd.read_sql(
        f"""
        SELECT to_char(data_chegada, 'DD/MM/YYYY') AS "Data",
               COUNT(*) AS "Total",
               SUM(CASE WHEN status='atendido' THEN 1 ELSE 0 END) AS "Atendidos",
               SUM(CASE WHEN status='desistiu' THEN 1 ELSE 0 END) AS "Desistências",
               ROUND(AVG(CASE WHEN data_atendimento IS NOT NULL
                   THEN EXTRACT(EPOCH FROM (data_atendimento - data_chegada)) / 60 END), 0
               ) AS "T. Médio (min)"
        FROM fila_espera
        WHERE data_chegada >= NOW() - INTERVAL '{periodo_dias} days'
        GROUP BY CAST(data_chegada AS DATE), to_char(data_chegada, 'DD/MM/YYYY')
        ORDER BY CAST(data_chegada AS DATE) DESC
        LIMIT 15
        """,
        engine,
    ).fillna(0)
    dados = [[str(v) for v in row] for _, row in df_pa.iterrows()]
    story.append(tabela_padrao(dados, list(df_pa.columns)))
    story.append(Spacer(1, 0.4*cm))

    # Seção 4: Estoque crítico
    story.append(Paragraph("4. Insumos em Situação Crítica", secao_style))
    df_est = pd.read_csv(config.DATA_DIR / "estoque_medicamentos.csv")
    df_crit = df_est[df_est["status_estoque"].isin(["esgotado","critico"])][
        ["nome","categoria","quantidade_atual","quantidade_minima","status_estoque"]
    ].rename(columns={
        "nome":"Insumo","categoria":"Categoria",
        "quantidade_atual":"Atual","quantidade_minima":"Mínimo","status_estoque":"Status"
    })
    dados = [[str(v) for v in row] for _, row in df_crit.iterrows()]
    if dados:
        story.append(tabela_padrao(dados, list(df_crit.columns)))
    else:
        story.append(Paragraph("Nenhum insumo em situação crítica no momento.", styles["Normal"]))

    doc.build(story)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/exportar/excel")
def exportar_excel(dias: int = Query(30, ge=7, le=365)):
    buf = _gerar_excel(dias)
    nome = f"relatorio_hospital_{date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome}"},
    )


@router.get("/exportar/pdf")
def exportar_pdf(dias: int = Query(30, ge=7, le=365)):
    buf = _gerar_pdf(dias)
    nome = f"relatorio_hospital_{date.today()}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nome}"},
    )
